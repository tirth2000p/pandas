
from flask import Flask, render_template, request, send_file
from werkzeug.utils import secure_filename
import datasc

from openpyxl import load_workbook, Workbook
from flask import Response
import io
buffer = io.BytesIO()

app = Flask(__name__)

out = 'out.xlsx'

@app.route('/upload')
def upload_file():
    return render_template('upload.html')


@app.route('/uploader', methods=['GET', 'POST'])
def uploader_file():
    if request.method == 'POST':
        f = request.files['file']
        f.save(f.filename)
        print(f.filename)
        df = datasc.data_ext(f.filename)
        df.to_excel(buffer,index=False)
        df.to_excel(out, index=False)

        book = load_workbook(out)
        sheet = book.active

        # load XLSX workbook
        wb = Workbook("out.xlsx")

        # save workbook as HTML file
        wb.save("templates/workbook.html")


        # headers = {
        #     'Content-Disposition': 'attachment; filename=output.xlsx',
        #     'Content-type': 'application/vnd.ms-excel'
        # }
        #
        # return Response(buffer.getvalue(), mimetype='application/vnd.ms-excel', headers=headers)
        return render_template('workbook.html', sheet=sheet)

@app.route('/download', methods=['GET', 'POST'])
def download_file():
    if request.method == 'GET':

        return send_file(out, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)